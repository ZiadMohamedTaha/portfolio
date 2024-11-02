import tkinter as tk

from datetime import datetime
from tracemalloc import start
import ZI  # Ensure ZI module is available
import pandas as pd 
from openpyxl import load_workbook
import os

class App:

    def __init__(self):
        self.file_path="D:\\project_covenant\\ورديات.xlsx"
        try:
            self.df = pd.read_excel(self.file_path,sheet_name='Sheet1')
        except FileNotFoundError:
            self.df = pd.DataFrame()  # إذا لم يكن الملف موجودًا، نبدأ بإطار بيانات فارغ

  
        # Create the main window
        self.main_window = tk.Tk()
        self.main_window.attributes('-fullscreen', True)
        self.main_window.overrideredirect(True)

        frame = tk.Frame(self.main_window, bg='gray', relief='raised', bd=2)
        frame.pack(fill='x')

        close_button = tk.Button(frame, text="X", command=self.close_app)
        close_button.pack(side='right')

        fullscreen_button = tk.Button(frame, text="[ ]", command=lambda: self.toggle_fullscreen(self.main_window))
        fullscreen_button.pack(side='right')

        minimize_button = tk.Button(frame, text="_", command=lambda: self.minimize_app(self.main_window))
        minimize_button.pack(side='right')

        self.take_back_id_card_button_therd_page = ZI.zi_button(self.main_window, 'تسليم ورديه', 'blue', self.send, fg='black')
        self.take_back_id_card_button_therd_page.place(x=650, y=300)

        self.take_back_id_card_button_therd_page = ZI.zi_button(self.main_window, 'استلام ورديه', 'red', self.recive, fg='black')
        self.take_back_id_card_button_therd_page.place(x=650, y=400)

        self.login_button = ZI.zi_button(self.main_window, "استلام بضاعه", "green", self.give)
        self.login_button.place(x=650, y=500)

        self.save_button_forth_page = ZI.zi_button(self.main_window, 'خروج بضاعه ', 'yellow', self.output, fg='purple')
        self.save_button_forth_page.place(x=650, y=600)

    def send(self):
        self.next_window = tk.Toplevel(self.main_window)
        self.next_window.attributes('-fullscreen', True)
        self.next_window.overrideredirect(True)

        frame = tk.Frame(self.next_window, bg='gray', relief='raised', bd=2)
        frame.pack(fill='x')

        close_button = tk.Button(frame, text="X", command=self.close_app)
        close_button.pack(side='right')

        fullscreen_button = tk.Button(frame, text="[ ]", command=lambda: self.toggle_fullscreen(self.next_window))
        fullscreen_button.pack(side='right')

        minimize_button = tk.Button(frame, text="_", command=lambda: self.minimize_app(self.next_window))
        minimize_button.pack(side='right')

        back_button_nx = tk.Button(frame, text="<-", command=lambda: self.back_nx(self.next_window))
        back_button_nx.pack(side='left')

        self.next_window_send = ZI.text(self.next_window, 'تسليم ورديه')
        self.next_window_send.place(x=550, y=50, width=400, height=30)

        next_window_user_name_label = ZI.text(self.next_window, 'ادخل إسمك ')
        next_window_user_name_label.place(x=550, y=350)

        self.next_window_user_name_enter = tk.Entry(self.next_window)
        self.next_window_user_name_enter.place(x=550, y=400, width=400, height=20)

        next_window_password_label = ZI.text(self.next_window, 'ادخل كلمة المرور خاصتك')
        next_window_password_label.place(x=550, y=500)
        self.next_window_password_enter = tk.Entry(self.next_window, show="*")
        self.next_window_password_enter.place(x=550, y=550, width=400, height=20)
        self.enter_us_ps = ZI.zi_button(self.next_window, "حفظ", 'green', self.th)
        self.enter_us_ps.place(x=550, y=600, width=400, height=30)

    def th(self):
        self.dec_user_pass = {
            'محمد وليد': '0223045',
            'زياد محمد': '0223042',
            'يوسف عباس': '0223050',
            'محمد طه': '1975'
        }

        user_name = self.next_window_user_name_enter.get()
        password = self.next_window_password_enter.get()

        if user_name in self.dec_user_pass and self.dec_user_pass[user_name] == password:
            self.user_save = user_name
            current_time = datetime.now()
            time = current_time.strftime("%I:%M:%S %p")
            date = datetime.today()
            date_d=date.date()
            data = {
                'name': [self.user_save],
                'send_recive rosary': ['تسليم ورديه'],
                'date': [date_d],
                'time': [time]
            }

            new_df = pd.DataFrame(data)

            # Merge the new data with existing data
            if not self.df.empty:
                df_combined = pd.concat([self.df, new_df], ignore_index=True)
            else:
                df_combined = new_df

            # Save to Excel file, handling existing sheets properly
            if os.path.exists(self.file_path):
                with pd.ExcelWriter(self.file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                    df_combined.to_excel(writer, sheet_name='Sheet1', index=False, header=False, startrow=writer.sheets['Sheet1'].max_row)
            else:
                df_combined.to_excel(self.file_path, sheet_name='Sheet1', index=False)
            ZI.msg('حفظ','تم حفظ تسليم الورديه بنجاح')
            self.next_window.withdraw()
        else:
            ZI.msg('خطأ', 'حدث خطأ يرجى اعادة كتابة الاسم و كلمة المرور')
    def recive(self):
        self.thierd_window = tk.Toplevel(self.main_window)
        self.thierd_window.attributes('-fullscreen', True)
        self.thierd_window.overrideredirect(True)

        frame = tk.Frame(self.thierd_window, bg='gray', relief='raised', bd=2)
        frame.pack(fill='x')

        close_button = tk.Button(frame, text="X", command=self.close_app)
        close_button.pack(side='right')

        fullscreen_button = tk.Button(frame, text="[ ]", command=lambda: self.toggle_fullscreen(self.thierd_window))
        fullscreen_button.pack(side='right')

        minimize_button = tk.Button(frame, text="_", command=lambda: self.minimize_app(self.thierd_window))
        minimize_button.pack(side='right')

        back_button = tk.Button(frame, text="<-", command=lambda: self.back(self.thierd_window))
        back_button.pack(side='left')

        self.thierd_window_recive = ZI.text(self.thierd_window, 'استلام ورديه')
        self.thierd_window_recive.place(x=550, y=50, width=400, height=30)

        thierd_window_user_name_label = ZI.text(self.thierd_window,'ادخل إسمك ')
        thierd_window_user_name_label.place(x=550, y=350)

        self.thierd_window_user_name_enter = tk.Entry(self.thierd_window)
        self.thierd_window_user_name_enter.place(x=550, y=400, width=400, height=20)

        thierd_window_password_label = ZI.text(self.thierd_window, ' ادخل كلمة المرور خاصتك ')
        thierd_window_password_label.place(x=550, y=500)
        self.thierd_window_password_enter = tk.Entry(self.thierd_window, show="*")
        self.thierd_window_password_enter.place(x=550, y=550, width=400, height=20)
        self.thierd_us_ps = ZI.zi_button(self.thierd_window, "حفظ", 'green', self.recive_th)
        self.thierd_us_ps.place(x=550, y=600, width=400, height=30)

    def recive_th(self):
        user_name_thierd = self.thierd_window_user_name_enter.get()
        password_thierd = self.thierd_window_password_enter.get()
        dec_user_pass1 = {
            'محمد وليد': '0223045',
            'زياد محمد': '0223042',
            'يوسف عباس': '0223050',
            'محمد طه': '1975'
        }
        
        if user_name_thierd in dec_user_pass1 and dec_user_pass1[user_name_thierd] == password_thierd:
            self.user_save_thierd = user_name_thierd
            current_time1 = datetime.now()
            time1 = current_time1.strftime("%I:%M:%S %p")
            date1 = datetime.today()
            date_d=date1.date()
            data1 = {
                'name': [self.user_save_thierd],
                'send_recive rosary': ['استلام ورديه'],
                'date': [date_d],
                'time': [time1]
            }
            print(data1)

            new_df1 = pd.DataFrame(data1)

            # Merge the new data with existing data
            if not self.df.empty:
                df_combined1 = pd.concat([self.df, new_df1], ignore_index=True)
            else:
                df_combined1 = new_df1
            
            # Save to Excel file, handling existing sheets properly
            if os.path.exists(self.file_path):
                with pd.ExcelWriter(self.file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                    df_combined1.to_excel(writer, sheet_name='Sheet1', index=False, header=False, startrow=writer.sheets['Sheet1'].max_row)
            else:
                df_combined1.to_excel(self.file_path, sheet_name='Sheet1', index=False)
            
            ZI.msg('حفظ','تم حفظ إستلام الورديه بنجاح')
            self.thierd_window.withdraw()
        else:
            ZI.msg('خطأ', 'حدث خطأ يرجى اعادة كتابة الاسم و كلمة المرور')
        
           
    def output(self):
        pass  

    def give(self):
        pass  

    def minimize_app(self, window):
        window.attributes('-fullscreen', False)
        window.overrideredirect(False)
        window.iconify()

    def toggle_fullscreen(self, window):
        if window.attributes('-fullscreen'):
            window.attributes('-fullscreen', False)
            window.overrideredirect(False)
        else:
            window.attributes('-fullscreen', True)
            window.overrideredirect(True)

    def back(self, window):
        if hasattr(self, 'thierd_window'):
            self.thierd_window.destroy()

    def back_nx(self, window):
        if hasattr(self, 'next_window'):
            self.next_window.destroy()

    def close_app(self):
        if hasattr(self, 'main_window'):
            self.main_window.destroy()
        if hasattr(self, 'next_window'):
            self.next_window.destroy()
        if hasattr(self, 'thierd_window'):
            self.thierd_window.destroy()

    def start(self):
        self.main_window.mainloop()


if __name__ == "__main__":
    app = App()
    app.start()
while app is start():
    app.start()
    if app.start()== ord('q'):
        break
